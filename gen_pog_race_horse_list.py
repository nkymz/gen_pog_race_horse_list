# -*- coding: utf-8 -*-

import os
import re
import datetime
import time
import sys

import openpyxl
import requests
from bs4 import BeautifulSoup

WEEKDAY = ["(月)", "(火)", "(水)", "(木)", "(金)", "(土)", "(日)"]

args = sys.argv
is_sp_reg = True if len(args) > 1 and args[1] == "sp" else False

mytoday = datetime.date.today()
mynow = datetime.datetime.today()

path = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"

wbpath = (path + "POG_HorseList.xlsx").replace("\\", "/")
htmlpath = (path + "PO_race_horse_list.html").replace("\\", "/")

wb = openpyxl.load_workbook(wbpath)
wshl = wb["POHorseList"]
wsSettings = wb["Settings"]

age = wsSettings["B1"].value
login_id = wsSettings["B2"].value
password = wsSettings["B3"].value

LOGIN_INFO = {
    'pid': 'login',
    'action': 'auth',
    'return_url2': '',
    'mem_tp': '',
    'login_id': login_id,
    'pswd': password,
    'auto_login': ''
}

mysession = requests.Session()
login_url = "https://regist.netkeiba.com/account/"
time.sleep(1)
mypost = mysession.post(login_url, data=LOGIN_INFO)


def get_stable_comment(horse_no, race_id):
    target_url = 'http://race.netkeiba.com/?pid=race_old&id=' + "c" + race_id + '&mode=comment'
    time.sleep(1)
    r = mysession.get(target_url)  # requestsを使って、webから取得
    soup = BeautifulSoup(r.content.decode("euc-jp", "ignore").encode("euc-jp"), 'lxml')  # 要素を抽出

    if not soup.find("div", class_="race_comment_box"):
        return ""

    stable_comment = ""
    stable_comment_row = soup.find("div", class_="race_comment_box").find("table").find_all("tr")[horse_no]
    stable_comment_columns = stable_comment_row.find_all("td")
    stable_comment += stable_comment_columns[3].text + "【" + stable_comment_columns[4].text + "】"

    return stable_comment


def get_predictions(horse_name, race_id):
    target_url = 'http://race.netkeiba.com/?pid=yoso&id=' + "c" + race_id
    time.sleep(1)
    r = mysession.get(target_url)  # requestsを使って、webから取得
    soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

    if not soup.find("div", id="race_main").find("table"):
        return ""

    prediction_header_text = [t.text for t in soup.find("div", id="race_main").find("table").find("tr").find_all("th")]
    hn_col_index = prediction_header_text.index("馬名")
    table_rows = soup.find("div", id="race_main").find("table").find_all("tr")
    horse_names = [t.find_all("td")[hn_col_index].text for i, t in enumerate(table_rows) if i > 0]
    horse_index = horse_names.index(horse_name) + 1
    predictions = soup.find("div", id="race_main").find("table").find_all("tr")[horse_index].find_all("td")
    if not predictions:
        return ""
    prediction_marks = ""
    if "\nCP予想\n" in prediction_header_text:
        range_max = prediction_header_text.index("\nCP予想\n")
    else:
        range_max = hn_col_index
    for i in range(2, range_max):
        prediction_marks += predictions[i].text.strip() if predictions[i].text.strip("\n") in "◎○▲☆△" else "－"
    return prediction_marks


def get_training_result(horse_no, race_id):
    target_url = 'http://race.netkeiba.com/?pid=race_old&id=' + "c" + race_id + '&mode=oikiri'
    time.sleep(1)
    r = mysession.get(target_url)  # requestsを使って、webから取得
    soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

    if not soup.find("div", id="race_main").find("table"):
        return "0000/00/00(火)", "", "", "", "", [], "", "", "", ""

    training_result_row = soup.find("div", id="race_main").find("table").find_all("tr")[horse_no]
    training_result_columns = training_result_row.find_all("td")
    training_date = training_result_columns[3].text
    if training_date.split("/")[0] == "0000":
        return "0000/00/00(火)", "", "", "", "", [], "", "", "", ""
    training_course = training_result_columns[4].text
    training_course_condition = training_result_columns[5].text
    training_jockey = training_result_columns[6].text
    training_time_list = [t.text for t in training_result_columns[7].find("ul").find_all("li")]
    training_result_texts_list = [t.text for t in training_result_columns[7].find_all("p")]
    training_position = training_result_columns[8].text
    training_stride = training_result_columns[9].text
    training_eval_text = training_result_columns[10].text
    training_eval_rank = training_result_columns[11].text

    return training_date, training_course, training_course_condition, training_jockey, training_time_list, \
        training_result_texts_list, training_position, training_stride, training_eval_text, training_eval_rank


xlrow = 2

while wshl.cell(row=xlrow, column=1).value is not None:
    # print(trow)

    horseNm = wshl.cell(row=xlrow, column=2).value
    horseNmOrgn = wshl.cell(row=xlrow, column=3).value

    isHorseNmDtrmnd = False

    if len(horseNm) < 6:
        isHorseNmDtrmnd = True
    elif horseNm[-5] != "の":
        isHorseNmDtrmnd = True

    if isHorseNmDtrmnd and horseNmOrgn is not None:
        xlrow += 1
        continue

    horseURLsp = wshl.cell(row=xlrow, column=5).value
    if horseURLsp is None:
        xlrow += 1
        continue

    time.sleep(1)
    r = mysession.get(horseURLsp)
    soup = BeautifulSoup(r.content, 'lxml')

    horseNmNew = soup.find("p", class_="Name").string
    horseNmOrgnNew = soup.find("th", string="馬名の意味").find_next().string

    wshl.cell(row=xlrow, column=2).value = horseNmNew
    if horseNmOrgnNew != "-":
        wshl.cell(row=xlrow, column=3).value = horseNmOrgnNew

    xlrow += 1

horse_list = [[cell.value for cell in row] for row in wshl["A2:F" + str(xlrow - 1)]]

race_horse_list = []

target_url = 'http://race.netkeiba.com/?rf=navi'
r = mysession.get(target_url)  # requestsを使って、webから取得
soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

DateList = soup.find('div', class_='DateList_Box')

for DateItem in DateList.find_all('a'):

    if not is_sp_reg and DateItem.get('href').split('=')[-1][0] not in 'c':
        continue
    elif is_sp_reg and DateItem.get('href').split('=')[-1][0] not in 'n':
        continue

    race_mmdd = DateItem.get('href').split('=')[-1][1:3] + "/" + DateItem.get('href').split('=')[-1][3:5]
    race_month = int(DateItem.get('href').split('=')[-1][1:3])
    race_day = int(DateItem.get('href').split('=')[-1][3:5])

    target_url = 'http://race.netkeiba.com' + DateItem.get('href')
    time.sleep(1)
    r = mysession.get(target_url)  # requestsを使って、webから取得
    soup = BeautifulSoup(r.text, 'lxml')  # r.contentだと文字化けする

    if not soup.find("table"):
        continue
    for i, scheduled_horse in enumerate(soup.find("table").find_all("tr")):
        if i == 0:
            continue

        element = scheduled_horse.find_all("td")

        horse_name = element[0].string
        horse_url = element[0].find("a").get("href")
        track = element[2].string[0:2]
        race_no = ("0" + element[2].string[2:])[-3:]
        race_name = element[3].string
        if len(race_name.split("(")) == 1:
            grade = "NG"
        elif race_name.split("(")[1].startswith("G"):
            grade = race_name.split("(")[1][0:2]
        else:
            grade = "NG"
        race_id = element[3].find("a").get("href").split("=")[-1][1:]
        race_year = int(race_id[0:4])
        race_date = race_id[0:4] + "/" + race_mmdd
        status = element[5].string

        owner = "金子真人HD"
        origin = "エラーエラーエラー"
        isSeal = False
        for horse in horse_list:
            if horse_name == horse[1]:
                owner = horse[0].strip()
                origin = horse[2]
                if horse[5] == "封印":
                    isSeal = True
                else:
                    isSeal = False

        race_url = 'http://race.netkeiba.com/?pid=race_old&id=' + "c" + race_id
        time.sleep(1)
        r = mysession.get(race_url)  # requestsを使って、webから取得
        soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

        h1_list = soup.find_all('h1')
        race_attrib_list = h1_list[1].find_all_next('p', limit=4)
        course = race_attrib_list[0].string.strip()
        race_time = race_attrib_list[1].string[-5:]
        weather = race_attrib_list[1].string.split("/")[0].split("：")[1]
        course_condition = race_attrib_list[1].string.split("/")[1].split("：")[1]
        race_cond1 = race_attrib_list[2].string
        race_cond2 = race_attrib_list[3].string

        horse_tag = soup.find("a", href=horse_url)
        if not horse_tag:
            continue
        horse_row = horse_tag.find_previous("tr")
        if not horse_row.find("td", class_="umaban"):
            horse_no = "00"
            box_no = "0"
        else:
            horse_no = horse_row.find("td", class_="umaban").string
            box_no = horse_row.find("td", class_=re.compile("^waku")).string
        if not horse_row.find_all('td', class_='txt_l', limit=2)[1].find('a'):
            jockey = None
        else:
            jockey = horse_row.find_all('td', class_='txt_l', limit=2)[1].find('a').string
        if not horse_row.find('td', class_='txt_r'):
            odds = None
            pop_rank = None
        else:
            odds = horse_row.find('td', class_='txt_r').string
            pop_rank = horse_row.find('td', class_='txt_r').find_next('td').string

        race_date2 = datetime.date(race_year, race_month, race_day)
        race_date = race_date + WEEKDAY[race_date2.weekday()]

        result, result_time, result_last3f = "00", "0", "0"
        result_url = None
        if race_date2 < mytoday or (race_date2 == mytoday and mynow.hour >= 17):
            result_url = race_url.replace("race_old", "race") + "&mode=result"
            time.sleep(1)
            r = mysession.get(result_url)  # requestsを使って、webから取得
            soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出
            horse_tag = soup.find("a", href=horse_url)
            horse_row = horse_tag.find_previous("tr")
            if horse_row.find("td", class_="result_rank").string:
                result = horse_row.find("td", class_="result_rank").string.zfill(2)
            else:
                result = "99"
            result_time = horse_row.find_all("td")[7].string
            result = horse_row.find_all("td")[8].string if result == "99" else result
            result_last3f = horse_row.find_all("td")[11].string

        training_date, training_course, training_course_condition, training_jockey, training_time_list, \
            training_result_texts_list, training_position, training_stride, training_eval_text, training_eval_rank \
            = get_training_result(int(horse_no), race_id)
        prediction_marks = get_predictions(horse_name, race_id)
        stable_comment = get_stable_comment(int(horse_no), race_id)

        sort_key = race_date + race_time + race_no + track + result + horse_no.zfill(2) + horse_name

        race_horse_list.append(
                [sort_key, race_date, race_time, track, race_no, race_name, grade, course, race_cond1, race_cond2,
                 horse_no, box_no, horse_name, jockey, odds, pop_rank, race_url, horse_url, owner, origin, result,
                 status, isSeal, result_url, training_date, training_course, training_course_condition, training_jockey,
                 training_time_list, training_result_texts_list, training_position, training_stride, training_eval_text,
                 training_eval_rank, prediction_marks, stable_comment, result_time, result_last3f, weather,
                 course_condition])

race_horse_list.sort()

f = open(htmlpath, mode="w", encoding="utf-8")

prev_date = None
prev_race_no = None
prev_race_time = None
prev_track = None

for i in race_horse_list:

    f.write("<!--" + str(i) + "-->\n")

    race_date = i[1]
    race_time = i[2]
    track = i[3]
    race_no = i[4]
    race_name = i[5]
    grade = i[6]
    course = i[7].replace("\xa0", " ")
    race_cond2 = i[9].replace("\xa0", " ")
    horse_no = i[10]
    box_no = i[11]
    horse_name = i[12]
    jockey = i[13]
    odds = i[14]
    pop_rank = i[15]
    race_url = i[16]
    horse_url = i[17]
    owner = i[18]
    origin = i[19]
    result = i[20]
    status = i[21]
    isSeal = i[22]
    result_url = i[23]
    training_date, training_course, training_course_condition, training_jockey, training_time_list, \
        training_result_texts_list, training_position, training_stride, training_eval_text, training_eval_rank, \
        prediction_marks, stable_comment, result_time, result_last3f, weather, course_condition \
        = i[24], i[25], i[26], i[27], i[28], i[29], i[30], i[31], i[32], i[33], i[34], i[35], i[36], i[37], i[38], i[39]

    sp = ' '
    
    if result != "00":
        race_url = result_url
        status = "【結果確定】"
    elif horse_no != "00":
        status = "【枠順確定】"
    elif status == "出走確定":
        status = "【出走確定】"
    elif not is_sp_reg:
        status = "【出走想定】"
    else:
        status = "【特別登録】"

    if box_no == "1":
        frame = '<span style="border: 1px solid; background-color:#ffffff; color:#000000;">1</span> '
    elif box_no == "2":
        frame = '<span style="border: 1px solid; background-color:#000000; color:#ffffff;">2</span> '
    elif box_no == "3":
        frame = '<span style="border: 1px solid; background-color:#ff0000; color:#ffffff;">3</span> '
    elif box_no == "4":
        frame = '<span style="border: 1px solid; background-color:#0000ff; color:#ffffff;">4</span> '
    elif box_no == "5":
        frame = '<span style="border: 1px solid; background-color:#ffff00; color:#000000;">5</span> '
    elif box_no == "6":
        frame = '<span style="border: 1px solid; background-color:#00ff00; color:#ffffff;">6</span> '
    elif box_no == "7":
        frame = '<span style="border: 1px solid; background-color:#ff8000; color:#000000;">7</span> '
    elif box_no == "8":
        frame = '<span style="border: 1px solid; background-color:#ff8080; color:#000000;">8</span> '
    else:
        frame = None

    s = '<h4>' + race_date + '</h4>\n'
    if prev_date is None:
        f.write(s)
    elif race_date != prev_date:
        f.write('</ul></li></ul>' + s)

    s = '<li> <a href="' + race_url + '">' + track + race_no + sp + race_name + status + '</a><br />\n'
    if weather == "&nbsp;":
        s2 = race_time + sp + course + sp + race_cond2 + '<br />\n<ul style="margin-left:-1em;">'
    else:
        s2 = race_time + sp + course + sp + race_cond2 + sp + weather + course_condition \
             + '<br />\n<ul style="margin-left:-1em;">'
    if prev_date is None or (prev_date is not None and race_date != prev_date):
        f.write('<ul style="margin-left:-1em;">' + s)
        f.write(s2)
    elif race_date + race_no + race_time + track != prev_date + prev_race_no + prev_race_time + prev_track:
        f.write('</ul></li>' + s)
        f.write(s2)

    if result == "01":
        s1 = '<li><span style="font-weight: 900; color:#FF0000;">1着</span>' + sp + frame + str(horse_no) + sp \
             + '<a href="' + horse_url + '">'
    elif result == "02" and grade != "NG":
        s1 = '<li><span style="font-weight: 700; color:#0000FF;">2着</span>' + sp + frame + str(horse_no) + sp \
             + '<a href="' + horse_url + '">'
    elif result in ["中止", "除外", "取消"]:
        s1 = "<li>" + result + sp + frame + str(horse_no) + sp + '<a href="' + horse_url + '">'
    elif result != "00":
        s1 = "<li>" + result.lstrip("0") + '着' + sp + frame + str(horse_no) + sp + '<a href="' + horse_url + '">'
    elif horse_no != "00":
        s1 = '<li>' + frame + str(horse_no) + sp + '<a href="' + horse_url + '">'
    else:
        s1 = '<li> <a href="' + horse_url + '">'
    if isSeal:
        s2 = '<s>' + horse_name + '</s>'
    else:
        s2 = horse_name
    s3 = sp + jockey if jockey else ""
    f.write(s1 + s2 + owner + '</a>' + s3 + '<br />\n')

    f.write(origin + '<br />\n')
    if result not in ["中止", "除外", "00"]:
        f.write(result_time + "(" + result_last3f + ")<br>\n")
    if odds is not None:
        f.write(str(odds) + '倍' + sp + str(pop_rank) + '番人気' + sp + prediction_marks + '<br />\n')
    if training_date[:4] != "0000":
        s = training_jockey + sp + training_date.split("/")[1] + "/" + training_date.split("/")[2].split("(")[0] + sp \
            + training_course + sp + training_course_condition + sp + training_stride + sp + "<br />\n"
        for t in training_time_list:
            s += t + " " if t != "-" else ""
        s += "[" + training_position + "]" + "<br />\n" if training_position else "<br />\n"
        for t in training_result_texts_list:
            s += t + sp
        s += "<br />\n" if training_result_texts_list else ""
        s += training_eval_text + training_eval_rank + "<br />\n"
        f.write(s)
    if stable_comment != "":
        f.write(stable_comment + "<br />\n")
    f.write('</li>\n')

    prev_date = race_date
    prev_race_no = race_no
    prev_race_time = race_time
    prev_track = track

s = str(mynow.year) + "/" + ("0" + str(mynow.month))[-2:] + "/" + ("0" + str(mynow.day))[-2:] \
    + WEEKDAY[mynow.weekday()] + " " + ("0" + str(mynow.hour))[-2:] + ":" + ("0" + str(mynow.minute))[-2:]
f.write('</ul></li></ul><p>' + s + ' 時点の情報より作成</p>\n')

f.close()

wb.save(wbpath)
wb.close()
